from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class AppStyles:
    def __init__(self):
        self.title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        self.accent_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.accent_fill = PatternFill(
            start_color="2B579A", end_color="2B579A", fill_type="solid"
        )
        self.bold_font = Font(name="Calibri", size=11, bold=True)
        self.regular_font = Font(name="Calibri", size=11)
