from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference


class DashboardBuilder:
    def __init__(self, wb, styles):
        self.wb = wb
        self.styles = styles

    def build(self):
        ws_app = self.wb.create_sheet(title="APP", index=0)
        ws_app.views.sheetView[0].showGridLines = True

        ws_app["B2"] = "SIMULADOR DE INVESTIMENTOS MULTICLASSE"
        ws_app["B2"].font = self.styles.title_font

        # 1. Configurações Globais
        ws_app["B9"] = "1. CONFIGURAÇÕES GLOBAIS"
        ws_app["B9"].font = self.styles.header_font
        ws_app["B9"].fill = self.styles.header_fill

        ws_app["B10"] = "Salário Base"
        ws_app["D10"] = 2000
        ws_app["D10"].number_format = "R$ #,##0.00"

        ws_app["B11"] = "Taxa Poupança Sugerida"
        ws_app["D11"] = 0.30
        ws_app["D11"].number_format = "0.0%"

        ws_app["B12"] = "Sugestão de Aporte Mensal (30%)"
        ws_app["D12"] = "=D10*D11"
        ws_app["D12"].number_format = "R$ #,##0.00"

        # 2. Seleção de Perfil e Aporte Real
        ws_app["B15"] = "2. SELEÇÃO DE PERFIL E APORTE"
        ws_app["B15"].font = self.styles.header_font
        ws_app["B15"].fill = self.styles.header_fill

        ws_app["B16"] = "Perfil de Investimento"
        ws_app["C16"] = "Moderado"
        ws_app["C16"].font = self.styles.bold_font

        # Validação de dados em C16 apontando para Database_Profiles
        dv_perfil = DataValidation(
            type="list", formula1="Database_Profiles!$F$2:$F$4", allow_blank=False
        )
        ws_app.add_data_validation(dv_perfil)
        dv_perfil.add("C16")

        ws_app["B17"] = "Valor Efetivo Aportado por Mês"
        ws_app["D17"] = "=D12"
        ws_app["D17"].number_format = "R$ #,##0.00"
        ws_app["D17"].font = self.styles.bold_font

        # Taxa de Retorno Média da Carteira (a.m.) em Português
        ws_app["B18"] = "Taxa de Retorno Média da Carteira (a.m.)"
        ws_app["D18"] = "=SOMARPRODUTO(C24:C29; E24:E29) / 12"
        ws_app["D18"].number_format = "0.00%"

        # 3. Alocação de Ativos por Perfil
        ws_app["B21"] = "3. DISTRIBUIÇÃO DA CARTEIRA POR ATIVO"
        ws_app["B21"].font = self.styles.header_font
        ws_app["B21"].fill = self.styles.header_fill

        headers_alloc = [
            "TIPO DE INVESTIMENTO",
            "Percentual na Carteira",
            "Aporte Alocado (R$)",
            "Retorno Esperado (a.a.)",
        ]
        for i, h in enumerate(headers_alloc, start=2):
            col_letter = get_column_letter(i)
            cell = ws_app[f"{col_letter}23"]
            cell.value = h
            cell.font = self.styles.accent_font
            cell.fill = self.styles.accent_fill

        inv_types = [
            "RENDA FIXA",
            "IMOBILIÁRIO",
            "MULTIMERCADO",
            "AÇÕES BR",
            "INTERNACIONAL",
            "CRIPTOMOEDAS",
        ]

        for idx, inv in enumerate(inv_types, start=24):
            ws_app[f"B{idx}"] = inv
            ws_app[f"B{idx}"].font = self.styles.regular_font

            # PROCV oficial em Português com ponto e vírgula e FALSO
            ws_app[f"C{idx}"] = (
                f'=PROCV($C$16 & "-" & B{idx}; Database_Profiles!$A$2:$E$19; 4; FALSO)'
            )
            ws_app[f"C{idx}"].number_format = "0.0%"

            # Valor em Reais alocado no mês
            ws_app[f"D{idx}"] = f"=C{idx}*$D$17"
            ws_app[f"D{idx}"].number_format = "R$ #,##0.00"

            # PROCV para Retorno Anual em Português
            ws_app[f"E{idx}"] = (
                f'=PROCV($C$16 & "-" & B{idx}; Database_Profiles!$A$2:$E$19; 5; FALSO)'
            )
            ws_app[f"E{idx}"].number_format = "0.0%"

        total_row = 24 + len(inv_types)
        ws_app[f"B{total_row}"] = "TOTAL"
        ws_app[f"B{total_row}"].font = self.styles.bold_font

        # SOMA em Português
        ws_app[f"C{total_row}"] = f"=SOMA(C24:C{total_row-1})"
        ws_app[f"C{total_row}"].font = self.styles.bold_font
        ws_app[f"C{total_row}"].number_format = "0.0%"

        ws_app[f"D{total_row}"] = f"=SOMA(D24:D{total_row-1})"
        ws_app[f"D{total_row}"].font = self.styles.bold_font
        ws_app[f"D{total_row}"].number_format = "R$ #,##0.00"

        # 4. Projeções de Cenários de Longo Prazo
        ws_app["B32"] = "4. PROJEÇÕES DE LONGO PRAZO (CENÁRIOS)"
        ws_app["B32"].font = self.styles.header_font
        ws_app["B32"].fill = self.styles.header_fill

        ws_app["B33"] = "Anos"
        ws_app["C33"] = "Total Aportado (Capital)"
        ws_app["D33"] = "Patrimônio Acumulado (com Juros)"
        ws_app["E33"] = "Rendimento Mensal Estimado"
        for col in ["B", "C", "D", "E"]:
            cell = ws_app[f"{col}33"]
            cell.font = self.styles.accent_font
            cell.fill = self.styles.accent_fill

        anos_lista = [2, 5, 10, 20, 30]
        for idx, anos in enumerate(anos_lista, start=34):
            ws_app[f"B{idx}"] = anos
            ws_app[f"C{idx}"] = f"=$D$17*{anos}*12"
            ws_app[f"C{idx}"].number_format = "R$ #,##0.00"

            # VF (Valor Futuro) em Português com ponto e vírgula
            ws_app[f"D{idx}"] = f"=VF($D$18; B{idx}*12; -$D$17)"
            ws_app[f"D{idx}"].number_format = "R$ #,##0.00"

            ws_app[f"E{idx}"] = f"=D{idx}*$D$18"
            ws_app[f"E{idx}"].number_format = "R$ #,##0.00"

        # Gráfico 1: Alocação de Ativos (Barra Vertical)
        chart_alloc = BarChart()
        chart_alloc.type = "col"
        chart_alloc.style = 10
        chart_alloc.title = "Distribuição da Carteira por Ativo"
        chart_alloc.y_axis.title = "Aporte Alocado (R$)"
        chart_alloc.x_axis.title = "Tipo de Investimento"

        data_alloc = Reference(ws_app, min_col=4, min_row=23, max_row=29)
        cats_alloc = Reference(ws_app, min_col=2, min_row=24, max_row=29)
        chart_alloc.add_data(data_alloc, titles_from_data=True)
        chart_alloc.set_categories(cats_alloc)
        chart_alloc.legend = None

        ws_app.add_chart(chart_alloc, "G23")

        # Gráfico 2: Projeção de Longo Prazo (Linha)
        chart_proj = LineChart()
        chart_proj.title = "Projeção de Patrimônio Acumulado (Longo Prazo)"
        chart_proj.style = 13
        chart_proj.y_axis.title = "Patrimônio (R$)"
        chart_proj.x_axis.title = "Anos"

        data_proj = Reference(ws_app, min_col=4, min_row=33, max_row=38)
        cats_proj = Reference(ws_app, min_col=2, min_row=34, max_row=38)
        chart_proj.add_data(data_proj, titles_from_data=True)
        chart_proj.set_categories(cats_proj)
        chart_proj.legend = None

        ws_app.add_chart(chart_proj, "G38")

        return ws_app
