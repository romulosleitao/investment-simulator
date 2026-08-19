import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# -------------------------------------------------------------
# ABA 2: Database_Profiles (Backend de Referência criado primeiro)
# -------------------------------------------------------------
ws_db = wb.active
ws_db.title = "Database_Profiles"
ws_db.views.sheetView[0].showGridLines = True

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
regular_font = Font(name="Calibri", size=11)

headers_db = ["CHAVE", "PERFIL", "TIPO DE INVESTIMENTO", "%"]
for col_idx, h in enumerate(headers_db, start=1):
    cell = ws_db.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill

data_db = [
    ("Conservador-RENDA FIXA", "Conservador", "RENDA FIXA", 0.50),
    ("Conservador-IMOBILIÁRIO", "Conservador", "IMOBILIÁRIO", 0.30),
    ("Conservador-MULTIMERCADO", "Conservador", "MULTIMERCADO", 0.10),
    ("Conservador-AÇÕES BR", "Conservador", "AÇÕES BR", 0.10),
    ("Conservador-INTERNACIONAL", "Conservador", "INTERNACIONAL", 0.00),
    ("Conservador-CRIPTOMOEDAS", "Conservador", "CRIPTOMOEDAS", 0.00),
    ("Moderado-RENDA FIXA", "Moderado", "RENDA FIXA", 0.35),
    ("Moderado-IMOBILIÁRIO", "Moderado", "IMOBILIÁRIO", 0.32),
    ("Moderado-MULTIMERCADO", "Moderado", "MULTIMERCADO", 0.08),
    ("Moderado-AÇÕES BR", "Moderado", "AÇÕES BR", 0.10),
    ("Moderado-INTERNACIONAL", "Moderado", "INTERNACIONAL", 0.10),
    ("Moderado-CRIPTOMOEDAS", "Moderado", "CRIPTOMOEDAS", 0.05),
    ("Agressivo-RENDA FIXA", "Agressivo", "RENDA FIXA", 0.10),
    ("Agressivo-IMOBILIÁRIO", "Agressivo", "IMOBILIÁRIO", 0.20),
    ("Agressivo-MULTIMERCADO", "Agressivo", "MULTIMERCADO", 0.05),
    ("Agressivo-AÇÕES BR", "Agressivo", "AÇÕES BR", 0.35),
    ("Agressivo-INTERNACIONAL", "Agressivo", "INTERNACIONAL", 0.20),
    ("Agressivo-CRIPTOMOEDAS", "Agressivo", "CRIPTOMOEDAS", 0.10),
]

for row_idx, row_data in enumerate(data_db, start=2):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws_db.cell(row=row_idx, column=col_idx, value=val)
        cell.font = regular_font
        if col_idx == 4:
            cell.number_format = "0.0%"

# Colocando os 3 perfis únicos em uma coluna de apoio oculta na aba de banco (ex: coluna F, linhas 2 a 4)
perfis_unicos = ["Conservador", "Moderado", "Agressivo"]
ws_db["F1"] = "Perfis_Unicos"
for idx, perfil in enumerate(perfis_unicos, start=2):
    ws_db[f"F{idx}"] = perfil


# -------------------------------------------------------------
# ABA 1: APP (Dashboard Principal)
# -------------------------------------------------------------
ws_app = wb.create_sheet(title="APP", index=0)
ws_app.views.sheetView[0].showGridLines = True

accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
accent_font = Font(name="Calibri", size=11, bold=True, color="000000")
bold_font = Font(name="Calibri", size=11, bold=True)

ws_app["B2"] = "SIMULADOR DE INVESTIMENTOS"
ws_app["B2"].font = Font(name="Calibri", size=16, bold=True, color="1F4E78")

# 1. Configurações Globais
ws_app["B9"] = "CONFIGURAÇÕES"
ws_app["B9"].font = header_font
ws_app["B9"].fill = header_fill

ws_app["B10"] = "Salário"
ws_app["D10"] = 2000
ws_app["D10"].number_format = "R$ #,##0.00"

ws_app["B11"] = "Rendimento Carteira"
ws_app["D11"] = 0.006
ws_app["D11"].number_format = "0.0%"

ws_app["B12"] = "Sugestão de Investimento (30%)"
ws_app["D12"] = "=D10*0.3"
ws_app["D12"].number_format = "R$ #,##0.00"

# 2. Simulador de Investimento Mensal
ws_app["B14"] = "INVESTIMENTO MENSAL"
ws_app["B14"].font = header_font
ws_app["B14"].fill = header_fill

ws_app["B15"] = "Quanto investir por mês ?"
ws_app["D15"] = 200
ws_app["D15"].number_format = "R$ #,##0.00"

ws_app["B16"] = "Por Quantos Anos ?"
ws_app["D16"] = 5

ws_app["B17"] = "Taxa de Rendimento mensal ?"
ws_app["D17"] = "=D11"
ws_app["D17"].number_format = "0.00%"

ws_app["B18"] = "Patrimônio acumulado ?"
ws_app["D18"] = "=VF(D17; D16*12; -D15)"
ws_app["D18"].number_format = "R$ #,##0.00"

ws_app["B19"] = "Retorno Mensal Estimado ?"
ws_app["D19"] = "=D18*D11"
ws_app["D19"].number_format = "R$ #,##0.00"

# 3. Tabela de Cenários
ws_app["B21"] = "Cenários (Projeção de Longo Prazo)"
ws_app["B21"].font = header_font
ws_app["B21"].fill = header_fill

ws_app["B22"] = "Anos"
ws_app["C22"] = "Patrimônio Acumulado"
ws_app["D22"] = "Retorno Mensal Estimado"
for col in ["B", "C", "D"]:
    cell = ws_app[f"{col}22"]
    cell.font = accent_font
    cell.fill = accent_fill

anos_lista = [2, 5, 10, 20, 30]
for idx, anos in enumerate(anos_lista, start=23):
    ws_app[f"B{idx}"] = anos
    ws_app[f"C{idx}"] = f"=VF($D$17; B{idx}*12; -$D$15)"
    ws_app[f"C{idx}"].number_format = "R$ #,##0.00"
    ws_app[f"D{idx}"] = f"=C{idx}*$D$17"
    ws_app[f"D{idx}"].number_format = "R$ #,##0.00"

# 4. Alocação de Ativos por Perfil com Validação Dinâmica puxando do Database_Profiles
ws_app["B29"] = "DISTRIBUIÇÃO DE CARTEIRA POR PERFIL"
ws_app["B29"].font = header_font
ws_app["B29"].fill = header_fill

ws_app["B30"] = "PERFIL SELECIONADO"
ws_app["C30"] = "Moderado"
ws_app["C30"].font = bold_font

# Validação de dados em C30 apontando para a lista de perfis na aba Database_Profiles ($F$2:$F$4)
dv_perfil = DataValidation(
    type="list", formula1="Database_Profiles!$F$2:$F$4", allow_blank=False
)
ws_app.add_data_validation(dv_perfil)
dv_perfil.add("C30")

ws_app["B31"] = "VALOR A SER INVESTIDO POR MÊS"
ws_app["C31"] = "=D15"
ws_app["C31"].font = bold_font
ws_app["C31"].number_format = "R$ #,##0.00"

headers_alloc = ["TIPO DE INVESTIMENTO", "Percentual Sugerido", "Valores (R$)"]
for i, h in enumerate(headers_alloc, start=2):
    col_letter = get_column_letter(i)
    cell = ws_app[f"{col_letter}33"]
    cell.value = h
    cell.font = accent_font
    cell.fill = accent_fill

inv_types = [
    "RENDA FIXA",
    "IMOBILIÁRIO",
    "MULTIMERCADO",
    "AÇÕES BR",
    "INTERNACIONAL",
    "CRIPTOMOEDAS",
]

for idx, inv in enumerate(inv_types, start=34):
    ws_app[f"B{idx}"] = inv
    ws_app[f"B{idx}"].font = regular_font

    # PROCX dinâmico buscando a chave combinada Perfil + Tipo na aba Database_Profiles
    ws_app[f"C{idx}"] = (
        f'=PROCX($C$30 & "-" & B{idx}; Database_Profiles!$A$2:$A$19; Database_Profiles!$D$2:$D$19; 0)'
    )
    ws_app[f"C{idx}"].number_format = "0.0%"

    ws_app[f"D{idx}"] = f"=C{idx}*$C$31"
    ws_app[f"D{idx}"].number_format = "R$ #,##0.00"

total_row = 34 + len(inv_types)
ws_app[f"B{total_row}"] = "TOTAL"
ws_app[f"B{total_row}"].font = bold_font
ws_app[f"C{total_row}"] = f"=SOMA(C34:C{total_row-1})"
ws_app[f"C{total_row}"].font = bold_font
ws_app[f"C{total_row}"].number_format = "0.0%"

ws_app[f"D{total_row}"] = f"=SOMA(D34:D{total_row-1})"
ws_app[f"D{total_row}"].font = bold_font
ws_app[f"D{total_row}"].number_format = "R$ #,##0.00"

# Ajustando larguras das colunas
for ws in [ws_app, ws_db]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save("investment_simulator.xlsx")
print(
    "Planilha gerada com validação de dados vinculada à aba Database_Profiles e PROCX!"
)
